#coding = utf-8
from openpyxl import load_workbook
from pause import Pause
import time
def getConditions():
    #定义conditions为list列表
    conditions = []
    try:
        #打开Excel表格
        wb = load_workbook('pauseAll.xlsx')
        #获取当前正在显示的sheet
        sheet = wb.active
        #定义一个数组,存储对应的参数
        array = ['id','operate']
        #定义行循环,循环第二列到 sheet.max+row 行
        for i in range(2,sheet.max_row + 1):
            #定义condition为字典数据类型
            condition = {}
            #使用enumerate定义列循环
            for j,arr in enumerate(array):
                #condition存入sheet.cell获取单元格内容
                condition[arr] = sheet.cell(i,j+1).value
            #使用append将condition添加到新的对象
            conditions.append(condition)

    except FileNotFoundError:
        print("pauseAll.xlsx文件不存在")

    return conditions

def run():
    conditions = getConditions()
    op = Pause()
    for condition in conditions:
        op.clickMyWechat(condition['id'])
        try:
            # 进行前置操作
            op.auto_ad_before(condition['operate'])
            for i in range(0,500):
                i += 1
                try:
                    try:
                        op.auto_ad(condition['operate'])
                        print('=======================' + condition['operate'] + '成功=======================')
                    except Exception:
                        pass
                    # 判断是否跳出循环
                    if op.validate_xpath():
                        break
                except Exception:
                    continue

            op.wait_text("小程序广告")
            op.find_link_text("小程序广告")

            time.sleep(1)
            # 进行前置操作
            op.auto_ad_before(condition['operate'])
            for i in range(0,500):
                i += 1
                try:
                    try:
                        op.auto_ad(condition['operate'])
                        print('======================='+ condition['operate'] +'成功=======================')
                    except Exception:
                        pass
                    # 判断是否跳出循环
                    if op.validate_xpath():
                        break
                except Exception:
                    continue

        except Exception:
            continue

    #定义循环结束页面
    print('=======================处理完成=======================')

if __name__ == '__main__':
    print("=======================欢迎使用南讯传媒微信广告服务商平台自动化 暂停||删除 脚本=======================")
    run()



















